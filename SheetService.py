import os
from openpyxl import load_workbook

diretorio = "C:\\EmailBot\\files\\PlanilhaFuturosFuncionarios.xlsx"

def fill_excel_worksheet(email_employee, name_employee):
    try:
        count = 1
        wb = load_workbook(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
        # wb = load_workbook(diretorio)
        ws = wb.worksheets[0]

        while ws["A" + str(count+1)].value != None and ws["B" + str(count+1)].value != None:
            count += 1

        ws["A" + str(count+1)] = name_employee
        ws["B" + str(count+1)] = email_employee

        wb.save(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
        # wb.save(diretorio)

    except:
        raise ValueError('Não foi possível preencher a planilha com os dados do empregado.')

def get_all_emails():
    wb = load_workbook(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
    # wb = load_workbook(diretorio)
    ws = wb.worksheets[0]

    emails = []
    for cell in ws['B']:
        if cell.value == 'Email' or cell.value == None:
            continue
        else:
            emails.append(cell.value)
    
    return emails

def change_excel_status(email, status):
    try:
        wb = load_workbook(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
        # wb = load_workbook(diretorio)
        ws = wb.worksheets[0]

        count = 1
        for i in get_all_emails():
            
            if i == email:
                ws["C" + str(count+1)] = status
            
            count += 1

        wb.save(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
        # wb.save(diretorio)
    
    except:
        raise ValueError("Não foi possível alterar o status no excel.")


def check_if_email_exists_in_worksheet(email):
    wb = load_workbook(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
    # wb = load_workbook(diretorio)
    ws = wb.worksheets[0]

    validate = False
    for cell in ws["B"]:
        if cell.value == email:
            validate = True 
    
    return validate

def is_already_send(email):
    wb = load_workbook(os.environ["DIRETORIO"] + "PlanilhaFuturosFuncionarios.xlsx")
    # wb = load_workbook(diretorio)
    ws = wb.worksheets[0]

    count = 1
    status = ''
    for cell in ws["B"]:
        if cell.value == email:
            status = ws["C" + str(count)].value
        
        count += 1        
    
    if status == "ENVIADO" or status == "Email inválido":
        return True

    return False
